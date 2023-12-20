from PIL import Image, ImageDraw, ImageFont, ImageFilter
from openpyxl import load_workbook
import os
import uuid
import qrcode
 
try:
    def create(name):
        num = uuid.uuid4()
        img = qrcode.make(num)
        img.save(f'qr_{num}.png')
        # code = int(281020230000) + int(num)
        # my_code = EAN13(str(code), writer=ImageWriter()) 
        # my_code.save(f"qr_{num}")
            # code128.image(num).save(f"qr_{num}.png")
            # code128.image(num).save("qr_.png")
        im1 = Image.open('cert.png')
        im2 = Image.open(f'qr_{num}.png')
        # im3 = Image.open('qr.png')
        im3 = im2.resize((400,400))
        # im2 = im3.resize((2937,1582))
        back_im = im1.copy()
        back_im.paste(im3, (340,840))
        # back_im.paste(im2, (15750, 4120))
        back_im.save(f'with_qr_{num}.png', quality=95)
        # back_im.save('with_qr.png', quality=95)

        # width, height = 19464, 6296
        color = 'rgb(0, 0, 0)'
        font1 = ImageFont.truetype('Montserrat-ExtraBold.ttf', size=120)
        font2 = ImageFont.truetype('Montserrat-ExtraBold.ttf', size=30)
        # (username_x, username_y) = (width+1931, height+293)

        image = Image.open(f'with_qr_{num}.png')
        # image = Image.open('with_qr.png')
        draw = ImageDraw.Draw(image)

        # _, _, w, h = draw.textbbox((1002,625), name, font=font1)
        draw.text((570,545), name, font=font1, fill=color)

        id_text = f"Certificate ID: {num}"
        draw.text((570,1230), id_text, font=font2, fill=color)
    #                  # campus_name = f"of branch {str.upper(df['Branch'][i])}, {str.lower(df['Year'][i])} Year (20{df['YearNo'][i]}-{df['YearNo'][i]+4}), {str.upper(df['Campus'][i])} who has actively taken participate"
        image.save(f'cert_{num}.png')
        # im5 = Image.open(f'cert_{num}.png')
        # im5 = im5.resize((4866,1574))
        # im5.save(f'Team_ticket_{num}.png')
            # image.save('fguys.png')
        os.remove(f"qr_{num}.png")
        # os.remove("qr.png")
        os.remove(f'with_qr_{num}.png')
            # os.remove(f'np_ticket_{num}.png')
            # os.remove('with_qr.png')
        print(num)
except Exception as e:
    print(e)

wb = load_workbook("sheet.xlsx")
ws = wb.active
column = ws["A"]
number = [column[x].value for x in range(len(column))]
column2 = ws["B"]
name = [column2[x].value for x in range(len(column2))]

for i in range(len(number)):
    if type(number[i]) == int:
        create(name[i].lower().capitalize())